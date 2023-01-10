from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException
from openpyxl import Workbook, load_workbook
import time
from copy import copy

wb  = Workbook()
sheet1 = wb.active
sheet1.title = '표준교육과정 과목 운영기관' #시트명
sheet1.cell(row=1, column=1).value = "학사명"
sheet1.cell(row=1, column=2).value = "전공명"
sheet1.cell(row=1, column=3).value = "과목명"
sheet1.cell(row=1, column=4).value = '전공구분'
sheet1.cell(row=1, column=5).value = '강의시간'
sheet1.cell(row=1, column=6).value = '실습시간'
sheet1.cell(row=1, column=7).value = '기관'
sheet1.cell(row=1, column=8).value = '지역'

driver = webdriver.Chrome()

kconti_url = 'https://www.cb.or.kr/creditbank/stdPro/nStdPro1_1.do'
driver.get(kconti_url)
driver.maximize_window()


row = 2
def work(index): # 전공하나하나 과목 얻어오기
    global row
    list = driver.find_element_by_css_selector('div.listDateWrap01')
    select_list = list.find_elements_by_css_selector('li')
    if select_list.__len__() > index:
        item_html = select_list[index] #해당하는 과목에 대한 li부분 가져오기
        jungong_flag_s = item_html.find_element_by_css_selector('em').text
        subject = item_html.find_element_by_css_selector('a').text
        lecture_time = item_html.find_elements_by_css_selector('span')[2].text
        practice_time = item_html.find_elements_by_css_selector('span')[3].text
        kigwan_button = item_html.find_elements_by_css_selector('a')[2]

        # sheet1.cell(row=row, column=1).value = haksa_name
        # sheet1.cell(row=row, column=2).value = major_name
        # sheet1.cell(row=row, column=3).value = subject
        # sheet1.cell(row=row, column=4).value = jungong_flag_s
        # sheet1.cell(row=row, column=5).value = lecture_time
        # sheet1.cell(row=row, column=6).value = practice_time
        kigwan_button.click()
        get_kigwan_list(haksa_name, major_name, subject, jungong_flag_s, lecture_time, practice_time)
        row = row + 1 # 과목 구분
        return select_list.__len__()

def get_kigwan_list(haksa_name, major_name, subject, jungong_flag_s, lecture_time, practice_time): # 해당하는 과목을 운영하는 기관 가져오기
    global row
    for page in range(1, 15):
        all_html = driver.find_element_by_css_selector('div.listDateWrap01')
        select_list = all_html.find_elements_by_css_selector('li')
        if select_list[0].text == '현재 운영중인 평가인정기관이 없습니다.':
            sheet1.cell(row=row, column=1).value = haksa_name  # 해당하는 과목과 각 기관과 지역 표시 추후 피벗으로 활용하기 위해
            sheet1.cell(row=row, column=2).value = major_name
            sheet1.cell(row=row, column=3).value = subject
            sheet1.cell(row=row, column=4).value = jungong_flag_s
            sheet1.cell(row=row, column=5).value = lecture_time
            sheet1.cell(row=row, column=6).value = practice_time
            sheet1.cell(row=row, column=7).value = "개설기관 없음"
            row += 1
            break
        # col = 7
        for item in select_list:
            ki_gwan = item.find_element_by_css_selector('a').text
            region = item.find_element_by_css_selector('em').text
            sheet1.cell(row=row, column=1).value = haksa_name # 해당하는 과목과 각 기관과 지역 표시 추후 피벗으로 활용하기 위해
            sheet1.cell(row=row, column=2).value = major_name
            sheet1.cell(row=row, column=3).value = subject
            sheet1.cell(row=row, column=4).value = jungong_flag_s
            sheet1.cell(row=row, column=5).value = lecture_time
            sheet1.cell(row=row, column=6).value = practice_time
            sheet1.cell(row=row, column=7).value = ki_gwan
            sheet1.cell(row=row, column=8).value = region
            row += 1
            # sheet1.cell(row=row, column=col).value = ki_gwan # 처음에는 해당하는 과목에 모든 기관을 표시
            # col = col + 1
        time.sleep(1)
        if page % 10 == 0:
            driver.find_element_by_link_text('다음 페이지로 이동').click()
        else:
            next = str(page + 1)
            try:
                driver.find_element_by_link_text(next).click()
            except NoSuchElementException:  # 마지막 다음 page는 없으므로 예외상황 발생되므로 ...^^ 꼼수
                break

if __name__ == "__main__":
    all_list = driver.find_element_by_css_selector('#contents > div.innerContView > div.stdProtResult')
    haksa_all = copy(all_list.find_elements_by_css_selector('h4'))  # 모든 학사 학사명을 얻기위해 1개씩
    haksa_cnt = len(haksa_all)
    major_all = copy(all_list.find_elements_by_css_selector('ul'))  # 모든 학사 비례 전공들
    major_list = all_list.find_elements_by_css_selector('li')  # 모든 학사 비례 전공들

    for i in range(haksa_cnt):
        major_html = major_all[i]
        major_list[i] = len(major_html.find_elements_by_css_selector('li'))

    for seq in range(haksa_cnt): # 전공클릭하여 정보얻고 back 했을때 웹정보를 잊어버리기때문에 강제적으로 url 정보 할당
        haksa_name =  haksa_all[seq].text
        out_flag = 0

        for major_seq in range(major_list[seq]):

            form = "#contents > div.innerContView > div.stdProtResult > div > ul:nth-child({}) > li:nth-child({}) > a"
            url = form.format(str((seq+1)*2), str(major_seq+1))
            search_button = driver.find_element_by_css_selector(url)
            if search_button.text == input_major_name:
                major_name = search_button.text
                out_flag = 1
                break
        if out_flag == 1:
            break

    search_button.click()  # 전공명 클릭하여 해당하는 과목들 가져오기
    subject_cnt = work(0)
    for index in range(subject_cnt):
        #break #
        if index != 0:
            search_button.click() # 전공명 클릭하여 해당하는 과목들 가져오기
            work(index) # 과목들에서 index에 해당하는 과목에 대한 훈련기관 가져오기

        driver.get(kconti_url)
        form = "#contents > div.innerContView > div.stdProtResult > div > ul:nth-child({}) > li:nth-child({}) > a"
        url = form.format(str((seq + 1) * 2), str(major_seq + 1))
        search_button = driver.find_element_by_css_selector(url)

    driver.quit()

    wb.save("./excel_folder/" + input_major_name + " 기관.xlsx")
