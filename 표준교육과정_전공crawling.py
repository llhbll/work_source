from selenium import webdriver

from openpyxl import Workbook, load_workbook

wb  = Workbook()
sheet1 = wb.active
sheet1.title = '표준교육과정 과목중심' #시트명
sheet1.cell(row=1, column=2).value = "과목명"
sheet1.cell(row=1, column=3).value = '전공구분'
sheet1.cell(row=1, column=4).value = '강의시간'
sheet1.cell(row=1, column=5).value = '실습시간'

driver = webdriver.Chrome()

kconti_url = 'https://www.cb.or.kr/creditbank/stdPro/nStdPro1_1.do'

driver.get(kconti_url)

search_button = driver.find_element_by_css_selector("#contents > div.innerContView > div.stdProtResult > div > ul:nth-child(20) > li:nth-child(1) > a")
# ul:nth-child(6) > li:nth-child(2) -> 경영학 // ul:nth-child(20) > li:nth-child(1) -> 국어국문학
major_name = search_button.text
search_button.click()


all_list = driver.find_element_by_css_selector('div.listDateWrap01')
select_list = all_list.find_elements_by_css_selector('li')

row = 2

for item in select_list:
    jungong_flag_s = item.find_element_by_css_selector('em').text
    subject = item.find_element_by_css_selector('a').text
    lecture_time = item.find_elements_by_css_selector('span')[2].text
    practice_time = item.find_elements_by_css_selector('span')[3].text

    sheet1.cell(row=row, column=2).value = subject
    sheet1.cell(row=row, column=3).value = jungong_flag_s
    sheet1.cell(row=row, column=4).value = lecture_time
    sheet1.cell(row=row, column=5).value = practice_time
    row = row + 1

driver.quit()
wb.save("./excel_folder/" + major_name + ".xlsx")







