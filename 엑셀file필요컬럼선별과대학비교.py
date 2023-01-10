import os
from glob import glob
from openpyxl import load_workbook
from openpyxl import Workbook

path = r'E:\기본자료\평생교육원정보\타대학현황\수입지출'
file_list = glob(r'E:\기본자료\평생교육원정보\타대학현황\수입지출\*수입지출 현황.xlsx')
# glob('*.xls')

# uni_dic = {'성신':[3, '성신여자대학교부설평생교육원'], '동덕':[4, '동덕여자대학교부설평생교육원'],
#            '덕성':[5, '덕성여자대학교부설평생교육원'], '서울':[6, '서울여자대학교부설평생교육원'],
#            '숙명':[7, '숙명여자대학교 미래교육원', '숙명여자대학교부설평생교육원'],
#            '이화':[8, '이화여자대학교부설평생교육원', '이화여자대학교 글로벌미래평생교육원'] }
# year_dic = {'2015':[5, '2015'], '2016':[6, '2016'], '2017':[7, '2017'], '2018':[8, '2018'], '2019':[9, '2019']} # row 행
# index
uni_dic = {'건국':[3, '건국대학교미래지식교육원'],
           '경기':[4, '경기대학교부설평생교육원'],
           '경희':[5, '경희대학교부설사회교육원', '경희대학교부설글로벌미래교육원'],
           '경희국제':[6, '경희대학교부설평생교육원(국제)'],
           '고려':[7, '고려대학교부설평생교육원'],
           '광운':[8, '광운대학교부설정보과학교육원'],
           '국민':[9, '국민대학교부설평생교육원'],
           '동국':[10, '동국대학교부설평생교육원', '동국대학교 부설 미래융합교육원'],
           '명지':[11, '명지대학교부설사회교육원', '명지대학교 미래교육원'],
           '삼육':[12, '삼육대학교 평생교육원'],
           '상명':[13, '상명대학교부설평생교육원'],
           '서강':[14, '서강대학교게임&평생교육원', '서강대학교부설평생교육원'],
           '서경':[15, '서경대학교 예술교육원', '서경대학교예술종합평생교육원'],
           '서울과기':[16, '서울과학기술대학교부설평생교육원'],
           '서울교육':[17, '서울교육대학교 평생교육원'],
           '시립':[18, '서울시립대학교 평생교육원'],
           '세종':[19, '세종대학교미래교육원', '세종대학교글로벌지식평생교육원'],
           '숭실':[20, '숭실대학교글로벌미래교육원', '숭실대학교부설평생교육원'],
           '연세':[21, '연세대학교 미래교육원', '연세대학교부설평생교육원'],
           '중앙':[22, '중앙대학교평생교육원(서울)'],
           '한성':[23, '한성대학교부설디자인아트교육원', '한성대학교부설디자인아트평생교육원'],
           '한양':[24, '한양대학교 부설 미래인재교육원', '한양대학교부설사회교육원'],
           '홍익':[25, '홍익대학교부설문화예술평생교육원']
}
year_dic = {'2015':[5, '2015'], '2016':[6, '2016'], '2017':[7, '2017'], '2018':[8, '2018'], '2019':[9, '2019']} # row 행

wb_new  = Workbook()
sheet1 = wb_new.active
sheet2 = wb_new.create_sheet('주요대학 수입지출 현황')

sheet1.title = '타대학 년도별 수입지출 현황' #시트명
sheet1.cell(row=1, column=1).value = "결산년도"
sheet1.cell(row=1, column=2).value = "기관유형"
sheet1.cell(row=1, column=3).value = '공시기관'
sheet1.cell(row=1, column=4).value = '수강료수입'
sheet1.cell(row=1, column=5).value = '국고보조금'
sheet1.cell(row=1, column=6).value = '기타수입'
sheet1.cell(row=1, column=7).value = '수입합계'
sheet1.cell(row=1, column=8).value = '인건비'
sheet1.cell(row=1, column=9).value = '관리운영비'
sheet1.cell(row=1, column=10).value = '연구학생경비'
sheet1.cell(row=1, column=11).value = '기타비용'
sheet1.cell(row=1, column=12).value = '지출합계'

# sheet2.cell(row=4, column=3).value = uni_dic['성신'][1] # 명칭
# sheet2.cell(row=4, column=4).value = uni_dic['동덕'][1]
# sheet2.cell(row=4, column=5).value = uni_dic['덕성'][1]
# sheet2.cell(row=4, column=6).value = uni_dic['서울'][1]
# sheet2.cell(row=4, column=7).value = uni_dic['숙명'][1]
# sheet2.cell(row=4, column=8).value = uni_dic['이화'][1]
sheet2.cell(row=4, column=3).value = uni_dic['건국'][1]
sheet2.cell(row=4, column=4).value = uni_dic['경기'][1]
sheet2.cell(row=4, column=5).value = uni_dic['경희'][1]
sheet2.cell(row=4, column=6).value = uni_dic['경희국제'][1]
sheet2.cell(row=4, column=7).value = uni_dic['고려'][1]
sheet2.cell(row=4, column=8).value = uni_dic['광운'][1]
sheet2.cell(row=4, column=9).value = uni_dic['국민'][1]
sheet2.cell(row=4, column=10).value = uni_dic['동국'][1]
sheet2.cell(row=4, column=11).value = uni_dic['명지'][1]
sheet2.cell(row=4, column=12).value = uni_dic['삼육'][1]
sheet2.cell(row=4, column=13).value = uni_dic['상명'][1]
sheet2.cell(row=4, column=14).value = uni_dic['서강'][1]
sheet2.cell(row=4, column=15).value = uni_dic['서경'][1]
sheet2.cell(row=4, column=16).value = uni_dic['서울과기'][1]
sheet2.cell(row=4, column=17).value = uni_dic['서울교육'][1]
sheet2.cell(row=4, column=18).value = uni_dic['시립'][1]
sheet2.cell(row=4, column=19).value = uni_dic['세종'][1]
sheet2.cell(row=4, column=20).value = uni_dic['숭실'][1]
sheet2.cell(row=4, column=21).value = uni_dic['연세'][1]
sheet2.cell(row=4, column=22).value = uni_dic['중앙'][1]
sheet2.cell(row=4, column=23).value = uni_dic['한성'][1]
sheet2.cell(row=4, column=24).value = uni_dic['한양'][1]
sheet2.cell(row=4, column=25).value = uni_dic['홍익'][1]

sheet2.cell(row=5, column=2).value = year_dic['2015'][1]
sheet2.cell(row=6, column=2).value = year_dic['2016'][1]
sheet2.cell(row=7, column=2).value = year_dic['2017'][1]
sheet2.cell(row=8, column=2).value = year_dic['2018'][1]
sheet2.cell(row=9, column=2).value = year_dic['2019'][1]

results = []

def pick_uni_ins(var_year, var_uni, var_in_sum): #file의 결산년도, 공시기관을 읽어와 딕셔너리의 결산년도와 대학평교를 비교하여 처리
    for uni_key, uni_data_list in uni_dic.items(): # uni_data_list = [7, '숙명여자대학교 미래교육원', '숙명여자대학교부설평생교육원']
        for uni_data in uni_data_list: # uni_data 차례대로 7, '숙명여자대학교 미래교육원', '숙명여자대학교부설평생교육원'
            if var_uni == uni_data:
                for year_key, year_data in year_dic.items(): # year_data = [5, "2015"]
                    if var_year == year_data[1]: # year_data 차례대로  5, "2015"
                        sheet2.cell(row=year_data[0], column=uni_data_list[0]).value = var_in_sum


for file_name_raw in file_list:

    file_name = file_name_raw
    wb = load_workbook(filename=file_name, data_only=True)
    ws = wb['결산(8월)']

    for row in ws.rows:
        result = []
        if row[0].value == '공시년도':
            continue
        result.append(str(int(row[0].value) -1)) #공시년도 -> -1년 하여 결산년도로 변경
        result.append(row[2].value) #기관유형
        result.append(row[3].value) #공시기관
        result.append(row[5].value) #수강료수입
        result.append(row[9].value) #국고보조금 수입
        result.append(row[11].value) #기타수입
        in_sum = float(row[5].value) + float(row[9].value) + float(row[11].value)
        result.append(in_sum)
        result.append(row[14].value) #인건비
        result.append(row[17].value) #관리운영비
        result.append(row[20].value) #연구학생경비
        result.append(row[26].value)  #기타비용
        out_sum = float(row[14].value) + float(row[17].value) + float(row[20].value) + float(row[26].value)
        result.append(out_sum)

        results.append(result)

        pick_uni_ins(str(int(row[0].value) -1), row[3].value, in_sum)
#        print(row[3].value)
#print(results)


for i in results:
    sheet1.append(i)

wb_new.save(r"E:\기본자료\평생교육원정보\타대학현황\수입지출\results.xlsx")