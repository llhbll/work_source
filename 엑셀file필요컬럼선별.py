import os
from glob import glob
from openpyxl import load_workbook
from openpyxl import Workbook


path = r'D:\lhb\study_source\data\타대학현황\수입지출\근거자료'
file_list = glob(r'D:\lhb\study_source\data\타대학현황\수입지출\근거자료\*수입지출 현황.xlsx')
# glob('*.xls')

wb_new  = Workbook()
sheet1 = wb_new.active
sheet1.title = '타대학 년도별 수입지출 현황' #시트명
sheet1.cell(row=1, column=1).value = "결산년도"
sheet1.cell(row=1, column=2).value = "기관유형"
sheet1.cell(row=1, column=3).value = '공시기관'
sheet1.cell(row=1, column=4).value = '수강료수입'
sheet1.cell(row=1, column=5).value = '국고보조금'
sheet1.cell(row=1, column=6).value = '기타수입'
sheet1.cell(row=1, column=7).value = '수입합계'
sheet1.cell(row=1, column=8).value = '인건비'
sheet1.cell(row=1, column=9).value = '관리운영비'
sheet1.cell(row=1, column=10).value = '연구학생경비'
sheet1.cell(row=1, column=11).value = '기타비용'
sheet1.cell(row=1, column=12).value = '지출합계'

results = []

for file_name_raw in file_list:

    file_name = file_name_raw
    wb = load_workbook(filename=file_name, data_only=True)
    ws = wb.get_sheet_by_name('결산(8월)')

    for row in ws.rows:
        result = []
        if row[0].value == '공시년도':
            continue
        result.append(str(int(row[0].value) -1)) #공시년도 -> -1년 하여 결산년도로 변경
        result.append(row[2].value) #기관유형
        result.append(row[3].value) #공시기관
        result.append(row[5].value) #수강료수입
        result.append(row[9].value) #국고보조금 수입
        result.append(row[11].value) #기타수입
        in_sum = float(row[5].value) + float(row[9].value) + float(row[11].value)
        result.append(str(in_sum))
        result.append(row[14].value) #인건비
        result.append(row[17].value) #관리운영비
        result.append(row[20].value) #연구학생경비
        result.append(row[26].value)  #기타비용
        out_sum = float(row[14].value) + float(row[17].value) + float(row[20].value) + float(row[26].value)
        result.append(str(out_sum))

        results.append(result)
#        print(row[3].value)
#print(results)


for i in results:
    sheet1.append(i)

wb_new.save(r"D:\lhb\study_source\data\타대학현황\수입지출\results.xlsx")