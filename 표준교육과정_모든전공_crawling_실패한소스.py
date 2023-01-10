from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException
from openpyxl import Workbook, load_workbook
import time
from copy import copy

wb  = Workbook()
sheet1 = wb.active
sheet1.title = '표준교육과정 과목중심' #시트명
sheet1.cell(row=1, column=2).value = "과목명"
sheet1.cell(row=1, column=3).value = '전공구분'
sheet1.cell(row=1, column=4).value = '강의시간'
sheet1.cell(row=1, column=5).value = '실습시간'

driver = webdriver.Chrome()

kconti_url = 'https://www.cb.or.kr/creditbank/stdPro/nStdPro1_1.do'
driver.get(kconti_url)
driver.maximize_window()

row = 2

def work(haksa_name, major_name): # 전공하나하나 과목 얻어오기
    global row
    list = driver.find_element_by_css_selector('div.listDateWrap01')
    select_list = list.find_elements_by_css_selector('li')
    for item in select_list:
        jungong_flag_s = item.find_element_by_css_selector('em').text
        subject = item.find_element_by_css_selector('a').text
        lecture_time = item.find_elements_by_css_selector('span')[2].text
        practice_time = item.find_elements_by_css_selector('span')[3].text

        sheet1.cell(row=row, column=1).value = haksa_name
        sheet1.cell(row=row, column=2).value = major_name
        sheet1.cell(row=row, column=3).value = subject
        sheet1.cell(row=row, column=4).value = jungong_flag_s
        sheet1.cell(row=row, column=5).value = lecture_time
        sheet1.cell(row=row, column=6).value = practice_time
        row = row + 1

    driver.back()
    time.sleep(2)

all_list = copy(driver.find_element_by_css_selector('#contents > div.innerContView > div.stdProtResult'))
haksa_all = copy(all_list.find_elements_by_css_selector('h4')) # 모든 학사 학사명을 얻기위해 1개씩
haksa_cnt = len(haksa_all)
major_all = copy(all_list.find_elements_by_css_selector('ul')) # 모든 학사 비례 전공들

for seq in range(haksa_cnt): # 전공클릭하여 정보얻고 back 했을때 웹정보를 잊어버리기때문에 강제적으로 url 정보 할당
    haksa_name = haksa_all[seq].text

    major_html = major_all[seq]
    major_list = copy(major_html.find_elements_by_css_selector('li'))
    for major_data in major_list:
        search_button = major_data.find_element_by_css_selector('a')
        major_name = search_button.text
        search_button.click()
        work(haksa_name, major_name)

driver.quit()

wb.save("./excel_folder/" + "모든전공.xlsx")
