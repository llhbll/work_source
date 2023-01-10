from openpyxl import Workbook, load_workbook
import sqlite3
import unicodedata

con = sqlite3.connect('.\data\hakjum.db')
cur = con.cursor()

wb  = load_workbook(r'C:\Users\sungshin\PycharmProjects\study_source\data\과목.xlsx')
sheet1 = wb['org']

row = 5

for row_data in sheet1.rows:
    ls_subject = sheet1.cell(row=row, column=2).value
    cur.execute('select 전공 from t_전공 where 과목 = (?)', (ls_subject, ))
    #cur.execute('select 기관 from t_학습자수 where 과목 = (?) and 년도 = 2019', (ls_subject,))
    #cur.execute('select a.기관, sum(a.인원) from t_학습자수 a, t_기관 b where a.과목 = (?) and a.기관 = b.기관 and b.대학여부 = '대학교'
    # and a.년도 = 2019 and b.기관유형 <> '원격교육' group by a.기관' ,
    # cur.execute('select a.기관, sum(a.인원) from t_학습자수 a, t_기관 b where a.년도 = 2019 and a.과목 = (?) and a.기관 = b.기관  '
    #             'and b.기관유형 = "원격교육" group '
    #             'by a.기관 order '
    #             'by sum(a.인원) desc' ,
    #             (ls_subject,))
    rows = cur.fetchall()
    col = 13
    cnt = 0
    for ls_tmp in rows:
        sheet1.cell(row=row, column=col).value = ls_tmp[0]
        # sheet1.cell(row=row, column=col+1).value = ls_tmp[1]
        col += 1
        # col += 2
        cnt += 1
    sheet1.cell(row=row, column=12).value = cnt
    row += 1

con.close()
wb.save(r'C:\Users\sungshin\PycharmProjects\study_source\data\과목.xlsx')
