from bs4 import BeautifulSoup as bs
from pprint import pprint
import requests, re, os
from urllib.request import  urlretrieve
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from xlsxwriter.utility import xl_col_to_name

wb = Workbook()
sheet1 = wb.active
sheet1.title = '네이버 웹툰 완결편'
sheet1.cell(row=1, column=1).value = '이미지'
sheet1.cell(row=1, column=2).value = '제목'
sheet1.cell(row=1, column=3).value = '평점'
sheet1.cell(row=1, column=4).value = '링크'

try:
    if not (os.path.isdir('image')):
        os.mkdir(os.path.join('image'))
    if not (os.path.isdir('excel_folder')):
        os.mkdir(os.path.join('excel_folder'))
except OSError as e:
    if e.errno != errno.EEXITS:
        print("폴더생성실패")
        exit()
url = 'https://comic.naver.com/webtoon/finish.nhn'
html = requests.get(url)
soup = bs(html.text, 'html.parser')
all_data = soup.find('div', {'class':'list_area'})

data_list = all_data.findAll('li')

col = 1
row = 2
for data in data_list:
    img_data = data.find('img')
    img_src = img_data['src']
    a_data = data.find('a')
    title = a_data['title']
    title = re.sub('[^0-9a-zA-Zㄱ-힗]', '', title)
    link = "https://comic.naver.com" + a_data['href']
    strong = data.find('strong').text

    urlretrieve(img_src, './image/'+title+'.gif')
    img_file = Image('./image/' + title + '.gif')
    # pprint(img_file)
    #cell = sheet1.cell(row=row, column=1)
    img_file.anchor = xl_col_to_name(col-1) +str(row)
    #pprint('A' + str(col))
    sheet1.add_image(img_file)
    sheet1.cell(row=row, column=2).value = title
    sheet1.cell(row=row, column=3).value = strong
    sheet1.cell(row=row, column=4).value = link
#    col = col + 1
    row = row + 1

wb.save("./excel_folder/webtoon.xlsx")

